import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.services.scoring import score_student

HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _stage_labels(activity: dict) -> dict[str, str]:
    return {s["id"]: s["label"] for s in activity["manifest"]["stages"]}


def build_gradebook(activity: dict, students: list[dict], responses: list[dict]) -> dict:
    """One row per student, one column per marked stage, plus totals.

    This is the sheet a teacher actually transfers to Schoology, so its rules
    matter: an ungraded stage is left **blank**, never 0. A blank says "not
    marked"; a 0 says "this student earned nothing", and a spreadsheet can't
    tell the reader which one you meant after the fact.
    """
    stages = activity["manifest"]["stages"]
    marked_stages = [s for s in stages if s.get("marks")]
    by_student: dict[str, list[dict]] = {}
    for r in responses:
        by_student.setdefault(r["student_id"], []).append(r)

    headers = ["Student", "Grade", "Section"]
    for s in marked_stages:
        headers.append(f"{s['label']} /{s['marks']:g}")
    headers += ["Total Awarded", "Out Of", "Pending Review", "Fully Graded"]

    rows = []
    for student in students:
        score = score_student(stages, by_student.get(student["id"], []))
        by_stage = {b["stage_id"]: b for b in score["stages"]}
        row = [student["name"], student.get("grade", ""), student.get("section", "")]
        for s in marked_stages:
            b = by_stage.get(s["id"], {})
            # Blank for anything nobody has marked yet, including stages the
            # student never reached -- whether that becomes a zero is the
            # teacher's decision, not ours.
            row.append(b.get("awarded") if b.get("status") != "not_answered" else None)
        row += [
            score["awarded_total"],
            score["max_score"],
            score["pending_review"] or None,
            "Yes" if score["fully_graded"] else "No",
        ]
        rows.append(row)

    return {"headers": headers, "rows": rows, "marked_stages": marked_stages}


def _summary_stats(students: list[dict], responses: list[dict]) -> dict:
    joined = len(students)
    responded = len({r["student_id"] for r in responses})
    graded = [r for r in responses if r["correct"] is not None]
    correct = sum(1 for r in graded if r["correct"])
    return {
        "joined": joined,
        "responded": responded,
        "participation_rate": round(100 * responded / joined) if joined else 0,
        "graded_responses": len(graded),
        "correct_responses": correct,
        "correct_rate": round(100 * correct / len(graded)) if graded else None,
    }


def build_csv(session: dict, activity: dict, students: list[dict], responses: list[dict]) -> str:
    students_by_id = {s["id"]: s for s in students}
    stage_labels = _stage_labels(activity)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Student Name", "Grade", "Section", "Stage", "Submitted At", "Correct",
        "Answer", "Auto Mark", "Teacher Mark",
    ])
    for r in responses:
        student = students_by_id.get(r["student_id"], {})
        writer.writerow([
            student.get("name", "Unknown"),
            student.get("grade", ""),
            student.get("section", ""),
            stage_labels.get(r.get("stage_id"), r.get("stage_id", "")),
            r["submitted_at"],
            r.get("correct", ""),
            r.get("answer", ""),
            # "" not 0 for an unmarked response, matching the Marks sheet.
            "" if r.get("mark") is None else r["mark"],
            "" if r.get("teacher_mark") is None else r["teacher_mark"],
        ])
    return output.getvalue()


def build_pdf(
    session: dict,
    activity: dict,
    students: list[dict],
    responses: list[dict],
    focus_violations: list[dict] | None = None,
) -> bytes:
    focus_violations = focus_violations or []
    students_by_id = {s["id"]: s for s in students}
    stage_labels = _stage_labels(activity)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()

    elements = [
        Paragraph("LISM AI Classroom Report", styles["Title"]),
        Paragraph(f"Activity: {activity['title']}", styles["Heading2"]),
        Paragraph(f"Session Code: {session['code']}", styles["Normal"]),
        Paragraph(f"Students joined: {len(students)} | Responses: {len(responses)}", styles["Normal"]),
        Spacer(1, 16),
    ]

    gradebook = build_gradebook(activity, students, responses)
    if gradebook["marked_stages"]:
        elements.append(Paragraph("Marks", styles["Heading2"]))
        marks_data = [gradebook["headers"]]
        for row in gradebook["rows"]:
            marks_data.append(["" if c is None else (f"{c:g}" if isinstance(c, (int, float)) else str(c)) for c in row])
        marks_table = Table(marks_data, repeatRows=1)
        marks_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0e7490")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#ecfeff")]),
        ]))
        elements.append(marks_table)
        elements.append(Paragraph(
            "A blank mark means not yet reviewed, not zero.",
            styles["Italic"],
        ))
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Responses", styles["Heading2"]))

    data = [["Student", "Grade", "Stage", "Correct", "Answer"]]
    for r in responses:
        student = students_by_id.get(r["student_id"], {})
        data.append([
            student.get("name", "Unknown"),
            str(student.get("grade", "")),
            stage_labels.get(r.get("stage_id"), r.get("stage_id", "")),
            "Yes" if r.get("correct") else ("No" if r.get("correct") is False else "-"),
            str(r.get("answer", ""))[:60],
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
    ]))
    elements.append(table)

    elements.append(Spacer(1, 24))
    elements.append(Paragraph("Focus Report", styles["Heading2"]))
    if not focus_violations:
        elements.append(Paragraph("No focus violations recorded for this session.", styles["Normal"]))
    else:
        focus_data = [["Student", "Exit #", "Type", "Time"]]
        for v in focus_violations:
            student = students_by_id.get(v["student_id"], {})
            focus_data.append([
                student.get("name", "Unknown"),
                str(v["violation_number"]),
                v["type"],
                v["occurred_at"],
            ])
        focus_table = Table(focus_data, repeatRows=1)
        focus_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7f1d1d")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fef2f2")]),
        ]))
        elements.append(focus_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def build_excel(
    session: dict,
    activity: dict,
    students: list[dict],
    responses: list[dict],
    focus_violations: list[dict] | None = None,
) -> bytes:
    focus_violations = focus_violations or []
    stage_labels = _stage_labels(activity)
    violation_counts: dict[str, int] = {}
    for v in focus_violations:
        violation_counts[v["student_id"]] = max(violation_counts.get(v["student_id"], 0), v["violation_number"])

    wb = Workbook()

    # --- Summary ---
    summary = wb.active
    summary.title = "Summary"
    stats = _summary_stats(students, responses)
    rows = [
        ("Activity", activity["title"]),
        ("Session Code", session["code"]),
        ("Session Type", session.get("session_type", "")),
        ("Students Joined", stats["joined"]),
        ("Students Responded", stats["responded"]),
        ("Participation Rate", f"{stats['participation_rate']}%"),
        ("Graded Responses", stats["graded_responses"]),
        ("Correct Responses", stats["correct_responses"]),
        ("Correct Rate", f"{stats['correct_rate']}%" if stats["correct_rate"] is not None else "-"),
        ("Total Focus Violations", len(focus_violations)),
        ("Students Locked (Focus)", sum(1 for c in violation_counts.values() if c >= 3)),
    ]
    for row in rows:
        summary.append(row)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 30
    for cell in summary["A"]:
        cell.font = Font(bold=True)

    # --- Marks ---
    # First sheet after Summary because it is the one that gets transferred to
    # the school gradebook; everything else is evidence behind it.
    gradebook = build_gradebook(activity, students, responses)
    if gradebook["marked_stages"]:
        marks_sheet = wb.create_sheet("Marks")
        marks_sheet.append(gradebook["headers"])
        for row in gradebook["rows"]:
            marks_sheet.append(row)
        _style_header_row(marks_sheet, len(gradebook["headers"]))
        marks_sheet.column_dimensions["A"].width = 24
        if any(r[-1] == "No" for r in gradebook["rows"]):
            marks_sheet.append([])
            marks_sheet.append([
                "Blank cells are not yet marked, not zero. "
                "Rows showing Fully Graded = No still need your review."
            ])

    # --- Responses ---
    resp_sheet = wb.create_sheet("Responses")
    headers = [
        "Student", "Grade", "Section", "Stage", "Correct", "Answer",
        "Auto Mark", "Teacher Mark", "Graded At", "Submitted At",
    ]
    resp_sheet.append(headers)
    students_by_id = {s["id"]: s for s in students}
    for r in responses:
        student = students_by_id.get(r["student_id"], {})
        resp_sheet.append([
            student.get("name", "Unknown"),
            student.get("grade", ""),
            student.get("section", ""),
            stage_labels.get(r.get("stage_id"), r.get("stage_id", "")),
            r.get("correct"),
            r.get("answer", ""),
            r.get("mark"),
            r.get("teacher_mark"),
            r.get("graded_at", ""),
            r["submitted_at"],
        ])
    _style_header_row(resp_sheet, len(headers))

    # --- Students ---
    student_sheet = wb.create_sheet("Students")
    headers = [
        "Name", "Grade", "Section", "Joined At", "Needs Help", "Help Requests",
        "Coach Messages", "Focus Violations", "Locked",
    ]
    student_sheet.append(headers)
    for s in students:
        violations = violation_counts.get(s["id"], 0)
        student_sheet.append([
            s["name"], s.get("grade", ""), s.get("section", ""), s["joined_at"],
            s.get("needs_help", False), s.get("help_requests", 0), s.get("coach_messages", 0),
            violations, violations >= 3,
        ])
    _style_header_row(student_sheet, len(headers))

    # --- Focus Report ---
    focus_sheet = wb.create_sheet("Focus Report")
    headers = ["Student", "Exit #", "Type", "Time"]
    focus_sheet.append(headers)
    for v in focus_violations:
        student = students_by_id.get(v["student_id"], {})
        focus_sheet.append([student.get("name", "Unknown"), v["violation_number"], v["type"], v["occurred_at"]])
    _style_header_row(focus_sheet, len(headers))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _style_header_row(sheet, num_columns: int) -> None:
    for col in range(1, num_columns + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
